"""Replace the day-calibrated loss impact functions with JRC depth-damage
curves in the Residential and Companies cost-benefit Entity files.

Discovery (Tasks 13-14): the impact functions produced by Task 9's
calibration (06b, following 06a's method) were fit against the 2024 event's
FLOOD DURATION hazard (days), but notebooks 11/12's cost-benefit uses the
JRC return-period FLOOD DEPTH hazard (meters). Since these are different
physical quantities, feeding the day-calibrated curve into the depth-based
hazard is not just a mislabeled unit — the curve's domain (0 to ~0.25 days
for Companies, 0 to ~28.75 "days" for Residential) doesn't correspond to
real depth values (0-7m), so CLIMADA extrapolates flat beyond the domain.
For Companies (whose curve saturates almost immediately) this made nearly
every flooded exposure take the maximum ~30% damage at nearly every return
period, inflating total climate risk to 901M USD (vs 320M for Schools, on
19x less exposed value).

Fix: the Schools depth-based curve already shipped in
entity_files_adaptation/..._Schools_calibrated_measures.xlsx (impf 301) is
*exactly* the JRC (Huizinga, De Moel & Szewczyk 2017, "Global flood
depth-damage functions", JRC105688) South America residential damage
function (Table 3-5). The same report's Table 3-9 gives an equivalent South
America commerce function. We use these two literature curves (residential
for Residential, commerce for Companies) in place of the day-calibrated
ones — matching, rather than inventing, the precedent already set for
Schools. Source: https://publications.jrc.ec.europa.eu/repository/bitstream/
JRC105688/global_flood_depth-damage_functions__10042017.pdf (Tables 3-5, 3-9).

This only touches the entity_files_adaptation/*_calibrated_measures.xlsx
copies used by notebooks 11/12 — entity_files_calibrated/ (Task 9's actual
deliverable, day-based, used by 06a/06b) is left untouched and remains
correct for its own (2024 event reconstruction) purpose.

Run from the notebooks/ directory with the climada_env environment.
"""
from pathlib import Path

import openpyxl

# Table 3-5 (residential) and Table 3-9 (commerce), JRC105688, South America.
# A near-zero point at 0.05m (mirroring the Schools curve's own 0/0.05 pair)
# avoids an instant jump right at zero depth.
JRC_SOUTH_AMERICA_CURVES = {
    "residential": [
        (0, 0.0), (0.05, 0.0), (0.5, 0.49), (1, 0.71), (1.5, 0.84),
        (2, 0.95), (3, 0.98), (4, 1.0), (5, 1.0), (6, 1.0),
    ],
    "commerce": [
        (0, 0.0), (0.05, 0.0), (0.5, 0.61), (1, 0.84), (1.5, 0.92),
        (2, 0.99), (3, 1.0), (4, 1.0), (5, 1.0), (6, 1.0),
    ],
}

CONFIGS = [
    {
        "entity_file": "Porto_Alegre_BRAZIL_Entity_Floods_Residential_calibrated_measures.xlsx",
        "impf_id": 101,
        "curve": "residential",
        "name": "Residential damage (JRC South America, Huizinga et al. 2017)",
    },
    {
        "entity_file": "Porto_Alegre_BRAZIL_Entity_Floods_Companies_calibrated_measures.xlsx",
        "impf_id": 201,
        "curve": "commerce",
        "name": "Company damage (JRC South America commerce, Huizinga et al. 2017)",
    },
]

ENTITY_DIR = Path("../data/entity_files_adaptation/")

for cfg in CONFIGS:
    path = ENTITY_DIR / cfg["entity_file"]
    wb = openpyxl.load_workbook(path)
    ws = wb["impact_functions"]
    header = [c.value for c in ws[1]]
    expected_header = ["impact_fun_id", "intensity", "mdd", "paa", "peril_id", "name", "intensity_unit"]
    assert header == expected_header, f"{cfg['entity_file']}: unexpected header {header}"

    id_col, intensity_col, mdd_col, paa_col, peril_col, name_col, unit_col = range(1, 8)

    # Find and remove all existing rows for this impf_id (the day-calibrated curve)
    rows_to_delete = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=id_col).value == cfg["impf_id"]
    ]
    assert rows_to_delete, f"{cfg['entity_file']}: no existing rows found for impf_id={cfg['impf_id']}"
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)

    # Append the JRC curve's rows at the end of the sheet
    curve = JRC_SOUTH_AMERICA_CURVES[cfg["curve"]]
    for intensity, mdd in curve:
        ws.append([cfg["impf_id"], intensity, mdd, 1, "FL", cfg["name"], "meter"])

    wb.save(path)
    print(f"{cfg['entity_file']}: replaced {len(rows_to_delete)} day-calibrated rows "
          f"with {len(curve)} JRC '{cfg['curve']}' rows for impf_id={cfg['impf_id']}")
