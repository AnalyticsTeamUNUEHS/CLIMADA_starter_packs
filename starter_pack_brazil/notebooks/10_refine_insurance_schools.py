"""Refine the insurance measure's attachment/cover from the model's own risk.

Computes the present-day unadapted exceedance curve for the Schools entity
(v2) against the JRC return-period hazard, then sets:
- risk transfer attachement = RP10 impact  (schools retain frequent losses)
- risk transfer cover       = RP100 impact (insurer pays up to the 1-in-100 loss)

Run from the notebooks/ directory with the climada_env environment.
"""
import copy
from pathlib import Path

import numpy as np
import openpyxl
from climada.engine import ImpactCalc
from climada.entity import Entity
from climada.hazard import Hazard

ENTITY_DIR = Path("../data/entity_files_adaptation/")
ENTITY_FILE = "Porto_Alegre_BRAZIL_Entity_Floods_Schools_calibrated_measures_v2.xlsx"
entity_path = ENTITY_DIR / ENTITY_FILE

HAZARD_DIR = Path("../data/hazard/")
HAZARD_RETURN_PERIODS = [10, 20, 50, 100, 200, 500]
haz_paths = [HAZARD_DIR / f"PortoAlegre_RP{rp}_depth.tif" for rp in HAZARD_RETURN_PERIODS]


def event_frequency_from_return_periods(return_periods):
    # Same conversion used in 09_costbenefit.ipynb
    rp_rev = copy.deepcopy(return_periods)[::-1]
    exceedance_frequency = np.array([1 / rp for rp in rp_rev])
    event_frequency = np.diff(np.concatenate([[0], exceedance_frequency]))
    return event_frequency[::-1]


entity = Entity.from_excel(entity_path)
entity.exposures.ref_year = 2025
entity.exposures.value_unit = "USD"

haz = Hazard.from_raster(
    files_intensity=haz_paths,
    attrs={
        "unit": "m",
        "event_name": [f"Return period {rp}" for rp in HAZARD_RETURN_PERIODS],
        "frequency": event_frequency_from_return_periods(HAZARD_RETURN_PERIODS),
    },
    haz_type="FL",
)

entity.exposures.assign_centroids(haz)

unadapted_impact = ImpactCalc(
    entity.exposures, entity.impact_funcs, haz
).impact(save_mat=False, assign_centroids=False)

freq_curve = unadapted_impact.calc_freq_curve(return_per=np.array([10, 100]))
rp10_impact, rp100_impact = freq_curve.impact

print(f"AAI (sanity check vs notebook): {unadapted_impact.aai_agg:,.2f} USD")
print(f"RP10 impact:  {rp10_impact:,.2f} USD")
print(f"RP100 impact: {rp100_impact:,.2f} USD")

assert rp100_impact > rp10_impact > 0, "Exceedance curve should be positive and increasing"

wb = openpyxl.load_workbook(entity_path)
ws = wb["measures"]
header = [c.value for c in ws[1]]
col_attach = header.index("risk transfer attachement") + 1
col_cover = header.index("risk transfer cover") + 1

for row_idx in range(2, ws.max_row + 1):
    if ws.cell(row=row_idx, column=1).value == "Portfolio flood insurance for schools":
        old_attach = ws.cell(row=row_idx, column=col_attach).value
        old_cover = ws.cell(row=row_idx, column=col_cover).value
        ws.cell(row=row_idx, column=col_attach, value=round(float(rp10_impact), 2))
        ws.cell(row=row_idx, column=col_cover, value=round(float(rp100_impact), 2))
        break
else:
    raise ValueError("Insurance measure row not found")

wb.save(entity_path)
print(f"attach: {old_attach:,} -> {rp10_impact:,.2f}")
print(f"cover:  {old_cover:,} -> {rp100_impact:,.2f}")
