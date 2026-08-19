"""Create the Companies Entity file with 3 adaptation measures.

Copies the calibrated Companies Entity (data/entity_files_calibrated/) into
data/entity_files_adaptation/ and appends 3 measures, mirroring the mechanism
mix used for Schools and Residential:

- Portable flood barriers for businesses (hazard linear transform, b=-0.3 m)
- Business interruption insurance        (risk transfer; attach/cover are
                                            first guesses, refined later)
- Relocation of high-risk businesses     (Exposures swap via the Task-10
                                            HDF5 pair)

Costs are first-guess estimates over the 9,742 businesses and 3.66bn USD
total value in this exposure set; refined later.

Run from the notebooks/ directory with the climada_env environment.
"""
import shutil
from pathlib import Path

import openpyxl

SRC_DIR = Path("../data/entity_files_calibrated/")
SRC = SRC_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Companies_calibrated.xlsx"

DST_DIR = Path("../data/entity_files_adaptation/")
DST_DIR.mkdir(exist_ok=True)
DST = DST_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Companies_calibrated_measures.xlsx"

shutil.copyfile(SRC, DST)

wb = openpyxl.load_workbook(DST)
ws = wb["measures"]

EXPECTED_HEADER = [
    "name", "color", "cost", "hazard intensity impact a", "hazard intensity impact b",
    "hazard high frequency cutoff", "hazard event set", "MDD impact a", "MDD impact b",
    "PAA impact a", "PAA impact b", "damagefunctions map", "assets file", "Region_ID",
    "risk transfer attachement", "risk transfer cover", "risk transfer cost factor",
    "peril_ID",
]
header = [c.value for c in ws[1]]
assert header == EXPECTED_HEADER, f"Unexpected measures header: {header}"

first_blank_row = None
for row_idx in range(2, ws.max_row + 2):
    if all(ws.cell(row=row_idx, column=c).value is None for c in range(1, len(EXPECTED_HEADER) + 1)):
        first_blank_row = row_idx
        break
assert first_blank_row is not None

existing_rows = [
    [ws.cell(row=r, column=c).value for c in range(1, len(EXPECTED_HEADER) + 1)]
    for r in range(2, first_blank_row)
]
# Same shipped no-op placeholder row as in the Residential file (see
# scratch_add_measures_residential.py) — overwrite it; fail loudly on
# anything else so we don't silently clobber a real measure.
PLACEHOLDER_ROW = ["measure", "0 0.4 0.4", 0, 1, 0, 0, "nil", 1, 0, 1, 0, "nil", "nil", 0, 0, 0, 1, "FL"]
assert existing_rows in ([], [PLACEHOLDER_ROW]), (
    f"Expected an empty measures sheet or the standard placeholder row, found: {existing_rows}"
)
first_blank_row = 2  # overwrite the placeholder (if present) instead of appending after it

# 9,742 businesses total (from the calibrated assets sheet's 'count' column)
N_BUSINESSES = 9_742
BARRIER_COST_PER_BUSINESS = 3_000  # USD, capital (portable flood barriers, door dams)
BARRIER_MAINTENANCE_RATE = 0.01  # per year
STUDY_YEARS = 25  # 2025-2050

barrier_capital = N_BUSINESSES * BARRIER_COST_PER_BUSINESS
barrier_maintenance = barrier_capital * BARRIER_MAINTENANCE_RATE * STUDY_YEARS
barrier_cost = round(barrier_capital + barrier_maintenance, 2)

RELOCATION_COST_PER_BUSINESS = 500_000  # USD, capital (rebuild/relocate 20 businesses)
N_RELOCATE = 20
RELOCATION_MAINTENANCE_PER_YEAR = 10_000  # USD/year, extra upkeep at new sites
relocation_capital = N_RELOCATE * RELOCATION_COST_PER_BUSINESS
relocation_maintenance = RELOCATION_MAINTENANCE_PER_YEAR * STUDY_YEARS
relocation_cost = round(relocation_capital + relocation_maintenance, 2)

new_rows = [
    (
        "Portable flood barriers for businesses", "0.9 0.3 0.1", barrier_cost,
        1, -0.3, 0, "nil",
        1, 0, 1, 0,
        "nil", "nil", 0,
        0, 0, 1, "FL",
    ),
    (
        "Business interruption insurance", "0.5 0.1 0.8", 2_000_000,
        1, 0, 0, "nil",
        1, 0, 1, 0,
        "nil", "nil", 0,
        1_000_000, 10_000_000, 1.5, "FL",
    ),
    (
        "Relocation of high-risk businesses", "0.8 0.0 0.2", relocation_cost,
        1, 0, 0, "nil",
        1, 0, 1, 0,
        "nil",
        "../data/entity_files_adaptation/Porto_Alegre_BRAZIL_Exposures_Companies_relocated.hdf5",
        0,
        0, 0, 1, "FL",
    ),
]

for offset, row in enumerate(new_rows):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=first_blank_row + offset, column=col_idx, value=value)

wb.save(DST)
print(f"Wrote {DST} with {len(new_rows)} measure rows")
print(f"Portable flood barriers cost: {barrier_cost:,.2f} USD "
      f"({barrier_capital:,.0f} capital + {barrier_maintenance:,.0f} maintenance)")
print(f"Relocation cost: {relocation_cost:,.2f} USD "
      f"({relocation_capital:,.0f} capital + {relocation_maintenance:,.0f} maintenance)")
