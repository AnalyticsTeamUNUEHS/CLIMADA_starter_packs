"""Create the v2 Entity file with 3 additional adaptation measures.

Copies the calibrated+measures Schools Entity and appends:
- School floor raising      (hazard linear transform, b=-0.4 m)
- Portfolio flood insurance (risk transfer; attach/cover are first guesses,
                             refined later from the model's exceedance curve)
- Partial relocation        (Exposures swap via the Task-3 HDF5)

Run from the notebooks/ directory with the climada_env environment.
"""
import shutil
from pathlib import Path

import openpyxl

ENTITY_DIR = Path("../data/entity_files_adaptation/")
SRC = ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Schools_calibrated_measures.xlsx"
DST = ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Schools_calibrated_measures_v2.xlsx"

shutil.copyfile(SRC, DST)

wb = openpyxl.load_workbook(DST)
ws = wb["measures"]

EXPECTED_HEADER = [
    "name", "color", "cost", "hazard intensity impact a", "hazard intensity impact b",
    "hazard high frequency cutoff", "hazard event set", "MDD impact a", "MDD impact b",
    "PAA impact a", "PAA impact b", "damagefunctions map", "assets file", "Region_ID",
    "risk transfer attachement", "risk transfer cover", "risk transfer cost factor",
    "peril_ID",
]
header = [c.value for c in ws[1]]
assert header == EXPECTED_HEADER, f"Unexpected measures header: {header}"

new_rows = [
    (
        "School floor raising", "0.9 0.5 0.1", 7_625_000,
        1, -0.4, 0, "nil",
        1, 0, 1, 0,
        "nil", "nil", 0,
        0, 0, 1, "FL",
    ),
    (
        "Portfolio flood insurance for schools", "0.3 0.1 0.7", 1_250_000,
        1, 0, 0, "nil",
        1, 0, 1, 0,
        "nil", "nil", 0,
        2_000_000, 15_000_000, 1.5, "FL",
    ),
    (
        "Partial relocation of high-risk schools", "0.9 0.1 0.1", 10_500_000,
        1, 0, 0, "nil",
        1, 0, 1, 0,
        "nil",
        "../data/entity_files_adaptation/Porto_Alegre_BRAZIL_Exposures_Schools_relocated.hdf5",
        0,
        0, 0, 1, "FL",
    ),
]

# Find the first blank row after the existing measure rows (file has trailing
# empty rows 5-9 with all-None cells)
first_blank_row = None
for row_idx in range(2, ws.max_row + 2):
    if all(ws.cell(row=row_idx, column=c).value is None for c in range(1, len(EXPECTED_HEADER) + 1)):
        first_blank_row = row_idx
        break
assert first_blank_row is not None

existing_names = [ws.cell(row=r, column=1).value for r in range(2, first_blank_row)]
assert existing_names == ["River dredging", "Rebuilding levees", "Early warning system"], existing_names

for offset, row in enumerate(new_rows):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=first_blank_row + offset, column=col_idx, value=value)

wb.save(DST)
print(f"Wrote {DST} with {len(new_rows)} new measure rows after {existing_names}")
