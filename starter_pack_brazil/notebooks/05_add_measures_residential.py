"""Create the Residential Entity file with 3 adaptation measures.

Copies the calibrated Residential Entity (data/entity_files_calibrated/) into
data/entity_files_adaptation/ and appends 3 measures, each a different
mechanism (mirrors the approach used for Schools):

- Home flood-proofing              (hazard linear transform, b=-0.3 m)
- Residential parametric insurance (risk transfer; attach/cover are first
                                     guesses, refined later from the model's
                                     exceedance curve)
- City-wide drainage upgrade       (hazard high-frequency cutoff at RP20)

Costs are first-guess engineering/programme estimates over the 762,239
buildings and 69.3bn USD total value in this exposure set; refined later.

Run from the notebooks/ directory with the climada_env environment.
"""
import shutil
from pathlib import Path

import openpyxl

SRC_DIR = Path("../data/entity_files_calibrated/")
SRC = SRC_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Residential_calibrated.xlsx"

DST_DIR = Path("../data/entity_files_adaptation/")
DST_DIR.mkdir(exist_ok=True)
DST = DST_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Residential_calibrated_measures.xlsx"

shutil.copyfile(SRC, DST)

wb = openpyxl.load_workbook(DST)
ws = wb["measures"]

EXPECTED_HEADER = [
    "name", "color", "cost", "hazard intensity impact a", "hazard intensity impact b",
    "hazard high frequency cutoff", "hazard event set", "MDD impact a", "MDD impact b",
    "PAA impact a", "PAA impact b", "damagefunctions map", "assets file", "Region_ID",
    "risk transfer attachement", "risk transfer cover", "risk transfer cost factor",
    "peril_ID",
]
header = [c.value for c in ws[1]]
assert header == EXPECTED_HEADER, f"Unexpected measures header: {header}"

# 762,239 buildings total (from the calibrated assets sheet's 'count' column)
N_BUILDINGS = 762_239
FLOODPROOFING_COST_PER_BUILDING = 800  # USD, capital
FLOODPROOFING_MAINTENANCE_RATE = 0.01  # per year
STUDY_YEARS = 25  # 2025-2050

floodproofing_capital = N_BUILDINGS * FLOODPROOFING_COST_PER_BUILDING
floodproofing_maintenance = floodproofing_capital * FLOODPROOFING_MAINTENANCE_RATE * STUDY_YEARS
floodproofing_cost = round(floodproofing_capital + floodproofing_maintenance, 2)

new_rows = [
    (
        "Home flood-proofing", "0.9 0.6 0.1", floodproofing_cost,
        1, -0.3, 0, "nil",
        1, 0, 1, 0,
        "nil", "nil", 0,
        0, 0, 1, "FL",
    ),
    (
        "Residential parametric insurance", "0.2 0.2 0.6", 5_000_000,
        1, 0, 0, "nil",
        1, 0, 1, 0,
        "nil", "nil", 0,
        5_000_000, 50_000_000, 1.5, "FL",
    ),
    (
        "City-wide drainage upgrade", "0.1 0.5 0.5", 300_000_000,
        1, 0, 1 / 20, "nil",
        1, 0, 1, 0,
        "nil", "nil", 0,
        0, 0, 1, "FL",
    ),
]

first_blank_row = None
for row_idx in range(2, ws.max_row + 2):
    if all(ws.cell(row=row_idx, column=c).value is None for c in range(1, len(EXPECTED_HEADER) + 1)):
        first_blank_row = row_idx
        break
assert first_blank_row is not None

existing_rows = [
    [ws.cell(row=r, column=c).value for c in range(1, len(EXPECTED_HEADER) + 1)]
    for r in range(2, first_blank_row)
]
# The shipped calibrated file carries a single no-op placeholder row
# (name='measure', cost=0, a=1/b=0/cutoff=0 — i.e. no effect on the risk
# calculation) rather than a truly empty sheet. Overwrite that placeholder;
# fail loudly on anything else so we don't silently clobber a real measure.
PLACEHOLDER_ROW = ["measure", "0 0.4 0.4", 0, 1, 0, 0, "nil", 1, 0, 1, 0, "nil", "nil", 0, 0, 0, 1, "FL"]
assert existing_rows in ([], [PLACEHOLDER_ROW]), (
    f"Expected an empty measures sheet or the standard placeholder row, found: {existing_rows}"
)
first_blank_row = 2  # overwrite the placeholder (if present) instead of appending after it

for offset, row in enumerate(new_rows):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=first_blank_row + offset, column=col_idx, value=value)

wb.save(DST)
print(f"Wrote {DST} with {len(new_rows)} measure rows")
print(f"Home flood-proofing cost: {floodproofing_cost:,.2f} USD "
      f"({floodproofing_capital:,.0f} capital + {floodproofing_maintenance:,.0f} maintenance)")
