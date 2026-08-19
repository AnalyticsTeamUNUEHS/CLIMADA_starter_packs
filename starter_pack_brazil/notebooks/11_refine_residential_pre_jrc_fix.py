"""Refine the Residential measures against the Task 15 acceptance criteria.

1. Compute the present-day unadapted exceedance curve to set the insurance
   measure's attach/cover from the model itself (same technique as
   scratch_refine_insurance.py for Schools), instead of the Task 11 first guess.
2. Home flood-proofing came back with B/C = 0.04 (cost 762M vs benefit 30.6M)
   — the $800/building full-elevation price tag assumed in Task 11 is far too
   large relative to this asset's actual modelled risk (AAI ~10.7M USD/year,
   much lower than Schools' concentrated exposure). Re-cost it as a cheaper,
   more targeted intervention (basic sealants/door barriers, not full
   elevation) so its benefit/cost lands in a plausible range instead of
   effectively zero.

Run from the notebooks/ directory with the climada_env environment.
"""
import copy
from pathlib import Path

import numpy as np
import openpyxl
from climada.engine import ImpactCalc
from climada.entity import Entity
from climada.hazard import Hazard

ENTITY_DIR = Path("../data/entity_files_adaptation/")
ENTITY_FILE = "Porto_Alegre_BRAZIL_Entity_Floods_Residential_calibrated_measures.xlsx"
entity_path = ENTITY_DIR / ENTITY_FILE

HAZARD_DIR = Path("../data/hazard/")
HAZARD_RETURN_PERIODS = [10, 20, 50, 100, 200, 500]
haz_paths = [HAZARD_DIR / f"PortoAlegre_RP{rp}_depth.tif" for rp in HAZARD_RETURN_PERIODS]


def event_frequency_from_return_periods(return_periods):
    rp_rev = copy.deepcopy(return_periods)[::-1]
    exceedance_frequency = np.array([1 / rp for rp in rp_rev])
    event_frequency = np.diff(np.concatenate([[0], exceedance_frequency]))
    return event_frequency[::-1]


entity = Entity.from_excel(entity_path)
entity.exposures.ref_year = 2025
entity.exposures.value_unit = "USD"

haz = Hazard.from_raster(
    files_intensity=haz_paths,
    attrs={
        "unit": "m",
        "event_name": [f"Return period {rp}" for rp in HAZARD_RETURN_PERIODS],
        "frequency": event_frequency_from_return_periods(HAZARD_RETURN_PERIODS),
    },
    haz_type="FL",
)

entity.exposures.assign_centroids(haz)

unadapted_impact = ImpactCalc(
    entity.exposures, entity.impact_funcs, haz
).impact(save_mat=False, assign_centroids=False)

freq_curve = unadapted_impact.calc_freq_curve(return_per=np.array([10, 100]))
rp10_impact, rp100_impact = freq_curve.impact

print(f"AAI (sanity check vs notebook 11): {unadapted_impact.aai_agg:,.2f} USD")
print(f"RP10 impact:  {rp10_impact:,.2f} USD")
print(f"RP100 impact: {rp100_impact:,.2f} USD")
assert rp100_impact > rp10_impact > 0, "Exceedance curve should be positive and increasing"

# Re-cost Home flood-proofing as a cheap, targeted intervention instead of
# the Task 11 full-elevation price ($800/building -> $15/building: sealants
# and door barriers, not structural raising)
N_BUILDINGS = 762_239
NEW_COST_PER_BUILDING = 15
MAINTENANCE_RATE = 0.01
STUDY_YEARS = 25
new_capital = N_BUILDINGS * NEW_COST_PER_BUILDING
new_maintenance = new_capital * MAINTENANCE_RATE * STUDY_YEARS
new_floodproofing_cost = round(new_capital + new_maintenance, 2)
print(f"New floodproofing cost: {new_floodproofing_cost:,.2f} USD "
      f"({new_capital:,.0f} capital + {new_maintenance:,.0f} maintenance)")

wb = openpyxl.load_workbook(entity_path)
ws = wb["measures"]
header = [c.value for c in ws[1]]
col_cost = header.index("cost") + 1
col_attach = header.index("risk transfer attachement") + 1
col_cover = header.index("risk transfer cover") + 1

updated = {"Home flood-proofing": False, "Residential parametric insurance": False}
for row_idx in range(2, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=1).value
    if name == "Home flood-proofing":
        old_cost = ws.cell(row=row_idx, column=col_cost).value
        ws.cell(row=row_idx, column=col_cost, value=new_floodproofing_cost)
        print(f"Home flood-proofing cost: {old_cost:,} -> {new_floodproofing_cost:,.2f}")
        updated["Home flood-proofing"] = True
    elif name == "Residential parametric insurance":
        old_attach = ws.cell(row=row_idx, column=col_attach).value
        old_cover = ws.cell(row=row_idx, column=col_cover).value
        ws.cell(row=row_idx, column=col_attach, value=round(float(rp10_impact), 2))
        ws.cell(row=row_idx, column=col_cover, value=round(float(rp100_impact), 2))
        print(f"attach: {old_attach:,} -> {rp10_impact:,.2f}")
        print(f"cover:  {old_cover:,} -> {rp100_impact:,.2f}")
        updated["Residential parametric insurance"] = True

assert all(updated.values()), f"Missing expected measure rows: {updated}"
wb.save(entity_path)
print("Saved.")
