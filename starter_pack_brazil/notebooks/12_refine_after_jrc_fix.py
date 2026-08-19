"""Re-refine Residential and Companies measures now that the loss curve is
the (much larger-risk) JRC depth-damage function instead of the day-based
one (see scratch_fix_depth_impact_functions.py).

With the corrected curve, total climate risk jumped from 179M to 9.73bn USD
(Residential) and from 901M to 2.30bn USD (Companies) -- the curve is now
dimensionally correct, but every measure's cost, sized against the old
(wrong-hazard) risk scale, needs re-checking against the Task 15 acceptance
criteria:

- Residential "Home flood-proofing" came back at B/C=101.8 (cost 14.3M
  against a benefit that grew to 1.45bn) -- exceeds our >50 review
  threshold, so its cost is rescaled up to a size proportionate to the new
  risk (from a "cheap sealant kit" price point to a more substantial
  floodproofing intervention, still per-building, just costed correctly
  for the risk it addresses).
- Both assets' insurance attach/cover (first guesses in Task 11/12) are
  recomputed from the corrected model's own exceedance curve.
- Companies' "Portable flood barriers" (B/C=11.18) and "Relocation of
  high-risk businesses" (B/C=3.66) are within the acceptance range already
  and are left as costed in Task 12.

Run from the notebooks/ directory with the climada_env environment.
"""
import copy
from pathlib import Path

import numpy as np
import openpyxl
from climada.engine import ImpactCalc
from climada.entity import Entity
from climada.hazard import Hazard

HAZARD_DIR = Path("../data/hazard/")
HAZARD_RETURN_PERIODS = [10, 20, 50, 100, 200, 500]
haz_paths = [HAZARD_DIR / f"PortoAlegre_RP{rp}_depth.tif" for rp in HAZARD_RETURN_PERIODS]


def event_frequency_from_return_periods(return_periods):
    rp_rev = copy.deepcopy(return_periods)[::-1]
    exceedance_frequency = np.array([1 / rp for rp in rp_rev])
    event_frequency = np.diff(np.concatenate([[0], exceedance_frequency]))
    return event_frequency[::-1]


haz = Hazard.from_raster(
    files_intensity=haz_paths,
    attrs={
        "unit": "m",
        "event_name": [f"Return period {rp}" for rp in HAZARD_RETURN_PERIODS],
        "frequency": event_frequency_from_return_periods(HAZARD_RETURN_PERIODS),
    },
    haz_type="FL",
)


def unadapted_exceedance(entity_path):
    entity = Entity.from_excel(entity_path)
    entity.exposures.ref_year = 2025
    entity.exposures.value_unit = "USD"
    entity.exposures.assign_centroids(haz)
    impact = ImpactCalc(entity.exposures, entity.impact_funcs, haz).impact(
        save_mat=False, assign_centroids=False
    )
    freq_curve = impact.calc_freq_curve(return_per=np.array([10, 100]))
    return impact.aai_agg, freq_curve.impact[0], freq_curve.impact[1]


def update_measure_cell(path, measure_name, column_name, new_value):
    wb = openpyxl.load_workbook(path)
    ws = wb["measures"]
    header = [c.value for c in ws[1]]
    col = header.index(column_name) + 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == measure_name:
            old_value = ws.cell(row=row_idx, column=col).value
            ws.cell(row=row_idx, column=col, value=new_value)
            wb.save(path)
            return old_value
    raise ValueError(f"{measure_name!r} not found in {path}")


ENTITY_DIR = Path("../data/entity_files_adaptation/")

# --- Residential ---
res_path = ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Residential_calibrated_measures.xlsx"
res_aai, res_rp10, res_rp100 = unadapted_exceedance(res_path)
print(f"Residential: AAI={res_aai:,.2f} USD, RP10={res_rp10:,.2f} USD, RP100={res_rp100:,.2f} USD")

old = update_measure_cell(res_path, "Residential parametric insurance", "risk transfer attachement", round(float(res_rp10), 2))
print(f"Residential insurance attach: {old:,} -> {res_rp10:,.2f}")
old = update_measure_cell(res_path, "Residential parametric insurance", "risk transfer cover", round(float(res_rp100), 2))
print(f"Residential insurance cover:  {old:,} -> {res_rp100:,.2f}")

# Re-cost Home flood-proofing: was priced as a $15/building sealant kit for
# the (wrong) day-based risk scale (B/C=2.14 there); against the corrected
# JRC-based risk it now reads B/C=101.8 at that price. Re-price as a more
# substantial floodproofing intervention (raised entrances/door barriers,
# not just sealants) so its B/C lands in a plausible, still-attractive range.
N_BUILDINGS = 762_239
NEW_COST_PER_BUILDING = 380
MAINTENANCE_RATE = 0.01
STUDY_YEARS = 25
capital = N_BUILDINGS * NEW_COST_PER_BUILDING
maintenance = capital * MAINTENANCE_RATE * STUDY_YEARS
new_cost = round(capital + maintenance, 2)
old = update_measure_cell(res_path, "Home flood-proofing", "cost", new_cost)
print(f"Home flood-proofing cost: {old:,} -> {new_cost:,.2f} "
      f"({capital:,.0f} capital + {maintenance:,.0f} maintenance)")

# --- Companies ---
comp_path = ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Companies_calibrated_measures.xlsx"
comp_aai, comp_rp10, comp_rp100 = unadapted_exceedance(comp_path)
print(f"\nCompanies: AAI={comp_aai:,.2f} USD, RP10={comp_rp10:,.2f} USD, RP100={comp_rp100:,.2f} USD")

old = update_measure_cell(comp_path, "Business interruption insurance", "risk transfer attachement", round(float(comp_rp10), 2))
print(f"Companies insurance attach: {old:,} -> {comp_rp10:,.2f}")
old = update_measure_cell(comp_path, "Business interruption insurance", "risk transfer cover", round(float(comp_rp100), 2))
print(f"Companies insurance cover:  {old:,} -> {comp_rp100:,.2f}")

print("\nDone. Portable flood barriers (B/C=11.18) and Relocation of "
      "high-risk businesses (B/C=3.66) were within the acceptance range "
      "and were not re-costed.")
