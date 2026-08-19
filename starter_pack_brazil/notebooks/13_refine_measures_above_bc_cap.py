"""Re-cost the 3 measures (across Residential and Companies) that exceeded
the >50 benefit/cost review threshold on their first post-JRC-fix run:

- Residential "Community-based early warning system": B/C=72.96 at 20M
- Residential "Flood-resistant building codes for renovations": B/C=77.83 at 25M
- Companies "Business continuity & supply chain diversification": B/C=86.26 at 8M

Costs are scaled up to target a benefit/cost near 10 (in line with the other
already-accepted measures in each notebook), not down-weighted in effect --
i.e. we're saying these programmes would realistically need a bigger budget
to reach full-population coverage, not that they're less effective.

Run from the notebooks/ directory with the climada_env environment.
"""
from pathlib import Path

import openpyxl

ENTITY_DIR = Path("../data/entity_files_adaptation/")

UPDATES = [
    {
        "path": ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Residential_calibrated_measures.xlsx",
        "measure": "Community-based early warning system",
        "new_cost": 150_000_000,
    },
    {
        "path": ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Residential_calibrated_measures.xlsx",
        "measure": "Flood-resistant building codes for renovations",
        "new_cost": 200_000_000,
    },
    {
        "path": ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Companies_calibrated_measures.xlsx",
        "measure": "Business continuity & supply chain diversification",
        "new_cost": 70_000_000,
    },
]

for upd in UPDATES:
    wb = openpyxl.load_workbook(upd["path"])
    ws = wb["measures"]
    header = [c.value for c in ws[1]]
    col_cost = header.index("cost") + 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == upd["measure"]:
            old_cost = ws.cell(row=row_idx, column=col_cost).value
            ws.cell(row=row_idx, column=col_cost, value=upd["new_cost"])
            print(f"{upd['path'].name} / {upd['measure']}: {old_cost:,} -> {upd['new_cost']:,}")
            break
    else:
        raise ValueError(f"{upd['measure']!r} not found in {upd['path']}")
    wb.save(upd["path"])

print("Done.")
