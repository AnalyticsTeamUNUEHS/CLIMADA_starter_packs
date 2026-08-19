"""Add 3 more measures to the Residential Entity file, bringing it to 6
total (matching the "3 mandatory + 3 additional for Brazil" rule already
applied to Schools). Each covers a mechanism not yet used by Residential's
first 3 measures (hazard transform, risk transfer, hazard frequency cutoff):

- Relocation of high-risk households  (Exposures swap, via Task 18's HDF5)
- Community-based early warning system (MDD linear transform)
- Flood-resistant building codes for renovations (PAA linear transform)

Run from the notebooks/ directory with the climada_env environment.
"""
from pathlib import Path

import openpyxl

ENTITY_DIR = Path("../data/entity_files_adaptation/")
PATH = ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Residential_calibrated_measures.xlsx"

wb = openpyxl.load_workbook(PATH)
ws = wb["measures"]

EXPECTED_HEADER = [
    "name", "color", "cost", "hazard intensity impact a", "hazard intensity impact b",
    "hazard high frequency cutoff", "hazard event set", "MDD impact a", "MDD impact b",
    "PAA impact a", "PAA impact b", "damagefunctions map", "assets file", "Region_ID",
    "risk transfer attachement", "risk transfer cover", "risk transfer cost factor",
    "peril_ID",
]
header = [c.value for c in ws[1]]
assert header == EXPECTED_HEADER, f"Unexpected measures header: {header}"

existing_names = [
    ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)
    if ws.cell(row=r, column=1).value is not None
]
assert existing_names == [
    "Home flood-proofing", "Residential parametric insurance", "City-wide drainage upgrade",
], f"Expected the 3 Task 11 measures already in place, found: {existing_names}"

# 762,239 buildings, 20,187 exposure points (from the calibrated assets sheet)
N_RELOCATE = 50
RELOCATION_COST_PER_HOUSEHOLD = 150_000  # USD, capital (rebuild/relocate)
RELOCATION_MAINTENANCE_PER_YEAR = 30_000  # USD/year, extra upkeep at new sites
STUDY_YEARS = 25
relocation_capital = N_RELOCATE * RELOCATION_COST_PER_HOUSEHOLD
relocation_cost = round(relocation_capital + RELOCATION_MAINTENANCE_PER_YEAR * STUDY_YEARS, 2)

new_rows = [
    (
        "Relocation of high-risk households", "0.2 0.7 0.3", relocation_cost,
        1, 0, 0, "nil",
        1, 0, 1, 0,
        "nil",
        "../data/entity_files_adaptation/Porto_Alegre_BRAZIL_Exposures_Residential_relocated.hdf5",
        0,
        0, 0, 1, "FL",
    ),
    (
        "Community-based early warning system", "0.4 0.8 0.2", 20_000_000,
        1, 0, 0, "nil",
        0.85, 0, 1, 0,
        "nil", "nil", 0,
        0, 0, 1, "FL",
    ),
    (
        "Flood-resistant building codes for renovations", "0.6 0.4 0.7", 25_000_000,
        1, 0, 0, "nil",
        1, 0, 0.8, 0,
        "nil", "nil", 0,
        0, 0, 1, "FL",
    ),
]

first_blank_row = None
for row_idx in range(2, ws.max_row + 2):
    if all(ws.cell(row=row_idx, column=c).value is None for c in range(1, len(EXPECTED_HEADER) + 1)):
        first_blank_row = row_idx
        break
assert first_blank_row is not None

for offset, row in enumerate(new_rows):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=first_blank_row + offset, column=col_idx, value=value)

wb.save(PATH)
print(f"Wrote {PATH} with {len(new_rows)} additional measure rows "
      f"(total now {len(existing_names) + len(new_rows)})")
print(f"Relocation cost: {relocation_cost:,.2f} USD "
      f"({relocation_capital:,.0f} capital + {RELOCATION_MAINTENANCE_PER_YEAR * STUDY_YEARS:,.0f} maintenance)")
