"""Recalibrate the JRC depth-damage curve amplitude for Residential and
Companies against the real 2024 observed loss, using RP200 depth as the
physical proxy for the 2024 event's severity.

Why RP200: the RP depth rasters and the 2024 duration raster come from
different hazard models/grids, so there is no direct depth footprint for
2024. Comparing flooded-area extent (km^2) across the two datasets shows
2024's extent (141.2 km^2) essentially matches RP200 (140.9 km^2), between
RP100 (134.3) and RP500 (148.0) -- consistent with 2024 being an
exceptional, near-record event for Porto Alegre.

The raw JRC curve applied to RP200 overestimates the real observed 2024
loss by ~20.7x (Residential) and ~2.8x (Companies). This script rescales
the MDD curve's amplitude (same shape/breakpoints, single multiplicative
factor) so it reproduces the real observed loss at that anchor, then
recomputes insurance attach/cover from the new (corrected) exceedance
curve -- mirroring 12_refine_after_jrc_fix.py's method, just anchored to
the right hazard this time. Other measures' "cost" cells are left
untouched: they represent programme/infrastructure budgets, not a
risk-scale-derived quantity, and rescaling them would just hide the real
finding (many measures are no longer cost-justified at the corrected,
much smaller risk).

Run from the notebooks/ directory with the climada_env environment.
"""
import copy
import shutil
from pathlib import Path

import numpy as np
import openpyxl
from climada.engine import ImpactCalc
from climada.entity import Entity, Exposures
from climada.hazard import Hazard

HAZARD_DIR = Path("../data/hazard/")
HAZARD_RETURN_PERIODS = [10, 20, 50, 100, 200, 500]
ENTITY_DIR = Path("../data/entity_files_adaptation/")
CALIB_DIR = Path("../data/entity_files_calibrated/")


def event_frequency_from_return_periods(return_periods):
    rp_rev = copy.deepcopy(return_periods)[::-1]
    exceedance_frequency = np.array([1 / rp for rp in rp_rev])
    event_frequency = np.diff(np.concatenate([[0], exceedance_frequency]))
    return event_frequency[::-1]


haz = Hazard.from_raster(
    files_intensity=[HAZARD_DIR / f"PortoAlegre_RP{rp}_depth.tif" for rp in HAZARD_RETURN_PERIODS],
    attrs={
        "unit": "m",
        "event_name": [f"Return period {rp}" for rp in HAZARD_RETURN_PERIODS],
        "frequency": event_frequency_from_return_periods(HAZARD_RETURN_PERIODS),
    },
    haz_type="FL",
)

haz_rp200 = Hazard.from_raster([HAZARD_DIR / "PortoAlegre_RP200_depth.tif"], haz_type="FL")
haz_rp200.event_id = np.array([1])
haz_rp200.frequency = np.array([1.0])
haz_rp200.date = np.array([1])


def rp200_modelled_loss(entity_path):
    entity = Entity.from_excel(entity_path)
    entity.exposures.assign_centroids(haz_rp200)
    imp = ImpactCalc(entity.exposures, entity.impact_funcs, haz_rp200).impact(
        save_mat=False, assign_centroids=False
    )
    return imp.at_event[0]


def rescale_impact_function(src_path, dst_path, impf_id, scale, new_name):
    shutil.copy(src_path, dst_path)
    wb = openpyxl.load_workbook(dst_path)
    ws = wb["impact_functions"]
    header = [c.value for c in ws[1]]
    col_id = header.index("impact_fun_id") + 1
    col_mdd = header.index("mdd") + 1
    col_name = header.index("name") + 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=col_id).value == impf_id:
            old_mdd = ws.cell(row=row_idx, column=col_mdd).value
            ws.cell(row=row_idx, column=col_mdd, value=round(old_mdd * scale, 6))
            ws.cell(row=row_idx, column=col_name, value=new_name)
    wb.save(dst_path)


def unadapted_exceedance(entity_path):
    entity = Entity.from_excel(entity_path)
    entity.exposures.ref_year = 2025
    entity.exposures.value_unit = "USD"
    entity.exposures.assign_centroids(haz)
    impact = ImpactCalc(entity.exposures, entity.impact_funcs, haz).impact(
        save_mat=False, assign_centroids=False
    )
    freq_curve = impact.calc_freq_curve(return_per=np.array([10, 100]))
    return impact.aai_agg, freq_curve.impact[0], freq_curve.impact[1]


def update_measure_cell(path, measure_name, column_name, new_value):
    wb = openpyxl.load_workbook(path)
    ws = wb["measures"]
    header = [c.value for c in ws[1]]
    col = header.index(column_name) + 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == measure_name:
            old_value = ws.cell(row=row_idx, column=col).value
            ws.cell(row=row_idx, column=col, value=new_value)
            wb.save(path)
            return old_value
    raise ValueError(f"{measure_name!r} not found in {path}")


ASSETS = [
    {
        "name": "Residential",
        "path": ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Residential_calibrated_measures.xlsx",
        "v2_path": ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Residential_calibrated_measures_v2.xlsx",
        "impf_id": 101,
        "observed_2024_loss": 212_170_680,
        "insurance_measure": "Residential parametric insurance",
        "new_impf_name": "Residential damage (JRC South America shape, Huizinga et al. 2017; amplitude rescaled to match observed 2024 loss via RP200 anchor)",
    },
    {
        "name": "Companies",
        "path": ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Companies_calibrated_measures.xlsx",
        "v2_path": ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Companies_calibrated_measures_v2.xlsx",
        "impf_id": 201,
        "observed_2024_loss": 438_755_400,
        "insurance_measure": "Business interruption insurance",
        "new_impf_name": "Company damage (JRC South America commerce shape, Huizinga et al. 2017; amplitude rescaled to match observed 2024 loss via RP200 anchor)",
    },
]

for a in ASSETS:
    unscaled_loss = rp200_modelled_loss(a["path"])
    k = a["observed_2024_loss"] / unscaled_loss
    print(f"\n=== {a['name']} ===")
    print(f"RP200 loss with raw JRC curve: {unscaled_loss:,.0f} USD")
    print(f"Observed 2024 loss:            {a['observed_2024_loss']:,.0f} USD")
    print(f"Rescale factor k = {k:.6f}")

    rescale_impact_function(a["path"], a["v2_path"], a["impf_id"], k, a["new_impf_name"])

    check_loss = rp200_modelled_loss(a["v2_path"])
    print(f"Sanity check -- RP200 loss with rescaled curve: {check_loss:,.0f} USD (should ~= observed)")

    aai, rp10, rp100 = unadapted_exceedance(a["v2_path"])
    print(f"New unadapted present-day AAI={aai:,.2f}  RP10={rp10:,.2f}  RP100={rp100:,.2f}")

    old_attach = update_measure_cell(a["v2_path"], a["insurance_measure"], "risk transfer attachement", round(float(rp10), 2))
    old_cover = update_measure_cell(a["v2_path"], a["insurance_measure"], "risk transfer cover", round(float(rp100), 2))
    print(f"Insurance attach: {old_attach:,.0f} -> {rp10:,.2f}")
    print(f"Insurance cover:  {old_cover:,.0f} -> {rp100:,.2f}")

print("\nDone. Other measures' cost cells left untouched -- see script docstring.")
