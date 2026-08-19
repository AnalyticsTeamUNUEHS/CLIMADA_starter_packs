"""Add 3 more measures to the Companies Entity file, bringing it to 6 total
(matching the "3 mandatory + 3 additional for Brazil" rule already applied
to Schools and Residential). Each covers a mechanism not yet used by
Companies' first 3 measures (hazard transform, risk transfer, exposures
swap):

- Levee protecting the commercial district (hazard high-frequency cutoff)
- Elevated inventory storage for retailers (MDD linear transform)
- Business continuity & supply chain diversification (PAA linear transform)

Run from the notebooks/ directory with the climada_env environment.
"""
from pathlib import Path

import openpyxl

ENTITY_DIR = Path("../data/entity_files_adaptation/")
PATH = ENTITY_DIR / "Porto_Alegre_BRAZIL_Entity_Floods_Companies_calibrated_measures.xlsx"

wb = openpyxl.load_workbook(PATH)
ws = wb["measures"]

EXPECTED_HEADER = [
    "name", "color", "cost", "hazard intensity impact a", "hazard intensity impact b",
    "hazard high frequency cutoff", "hazard event set", "MDD impact a", "MDD impact b",
    "PAA impact a", "PAA impact b", "damagefunctions map", "assets file", "Region_ID",
    "risk transfer attachement", "risk transfer cover", "risk transfer cost factor",
    "peril_ID",
]
header = [c.value for c in ws[1]]
assert header == EXPECTED_HEADER, f"Unexpected measures header: {header}"

existing_names = [
    ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)
    if ws.cell(row=r, column=1).value is not None
]
assert existing_names == [
    "Portable flood barriers for businesses", "Business interruption insurance",
    "Relocation of high-risk businesses",
], f"Expected the 3 Task 12 measures already in place, found: {existing_names}"

# 9,742 businesses (from the calibrated assets sheet's 'count' column)
N_BUSINESSES = 9_742
INVENTORY_COST_PER_BUSINESS = 1_500  # USD, capital (elevated racking/storage)
INVENTORY_MAINTENANCE_RATE = 0.01  # per year
STUDY_YEARS = 25
inventory_capital = N_BUSINESSES * INVENTORY_COST_PER_BUSINESS
inventory_cost = round(inventory_capital + inventory_capital * INVENTORY_MAINTENANCE_RATE * STUDY_YEARS, 2)

new_rows = [
    (
        "Levee protecting the commercial district", "0.1 0.6 0.3", 50_000_000,
        1, 0, 1 / 20, "nil",
        1, 0, 1, 0,
        "nil", "nil", 0,
        0, 0, 1, "FL",
    ),
    (
        "Elevated inventory storage for retailers", "0.7 0.4 0.1", inventory_cost,
        1, 0, 0, "nil",
        0.8, 0, 1, 0,
        "nil", "nil", 0,
        0, 0, 1, "FL",
    ),
    (
        "Business continuity & supply chain diversification", "0.3 0.6 0.6", 8_000_000,
        1, 0, 0, "nil",
        1, 0, 0.7, 0,
        "nil", "nil", 0,
        0, 0, 1, "FL",
    ),
]

first_blank_row = None
for row_idx in range(2, ws.max_row + 2):
    if all(ws.cell(row=row_idx, column=c).value is None for c in range(1, len(EXPECTED_HEADER) + 1)):
        first_blank_row = row_idx
        break
assert first_blank_row is not None

for offset, row in enumerate(new_rows):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=first_blank_row + offset, column=col_idx, value=value)

wb.save(PATH)
print(f"Wrote {PATH} with {len(new_rows)} additional measure rows "
      f"(total now {len(existing_names) + len(new_rows)})")
print(f"Elevated inventory storage cost: {inventory_cost:,.2f} USD")
